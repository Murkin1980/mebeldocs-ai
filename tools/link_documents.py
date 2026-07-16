#!/usr/bin/env python3
"""Index local accounting workbooks and propose evidence-based links to active outgoing ESF."""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "working" / "links"
ESF = ROOT / "data" / "working" / "esf" / "records.json"
OWN_TIN = "910226302322"


def clean(v) -> str:
    return " ".join(str(v or "").replace("\n", " ").split())


def doc_type(path: Path) -> str | None:
    s = str(path).casefold()
    if "счет на оплат" in s or "счёт на оплат" in s:
        return "invoice"
    if "накладн" in s or path.name.casefold().startswith("продажа"):
        return "goods_issue"
    if "авр" in s or "акт выполн" in s:
        return "service_act"
    return None


def parse_number(path: Path, text: str) -> str:
    patterns = [r"(?:счет|счёт|продажа|документа|акт)[^№]{0,30}№?\s*0*(\d+)", r"№\s*0*(\d+)"]
    for source in (path.stem, text[:3000]):
        for pattern in patterns:
            m = re.search(pattern, source, re.I)
            if m:
                return m.group(1)
    return ""


MONTHS = {"января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12}


def parse_date(path: Path, text: str) -> str:
    for source in (path.stem, text[:3000]):
        m = re.search(r"(?:от\s*)?(\d{1,2})[.,-](\d{1,2})[.,-](20\d{2})", source, re.I)
        if m:
            try: return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).strftime("%d.%m.%Y")
            except ValueError: pass
        m = re.search(r"(?:от\s*)?(\d{1,2})\s+(" + "|".join(MONTHS) + r")\s+(20\d{2})", source, re.I)
        if m:
            return datetime(int(m.group(3)), MONTHS[m.group(2).casefold()], int(m.group(1))).strftime("%d.%m.%Y")
    return ""


def decimal(v) -> Decimal | None:
    if isinstance(v, (int, float)):
        try: return Decimal(str(v)).quantize(Decimal("0.01"))
        except InvalidOperation: return None
    s = clean(v).replace(" ", "").replace(",", ".")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", s):
        try: return Decimal(s).quantize(Decimal("0.01"))
        except InvalidOperation: return None
    return None


def parse_workbook(path: Path) -> dict | None:
    kind = doc_type(path)
    if not kind or path.name.startswith("~$"):
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    flat = []
    rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            vals = [c.value for c in row]
            rows.append(vals)
            flat.extend(vals)
    wb.close()
    text = " | ".join(clean(v) for v in flat if clean(v))
    tins = sorted(set(re.findall(r"(?<!\d)\d{12}(?!\d)", text)))
    counterparties = [x for x in tins if x != OWN_TIN]
    totals = []
    for row in rows:
        labels = [clean(v).casefold().rstrip(":") for v in row]
        if "итого" in labels:
            idx = labels.index("итого")
            nums = [decimal(v) for v in row[idx + 1:]]
            nums = [n for n in nums if n is not None and n >= 0]
            if nums: totals.append(max(nums))
    if not totals:
        for m in re.finditer(r"(?:итого|на сумму)\D{0,30}([\d ]+[,.]\d{2}|\d{3,})", text, re.I):
            n = decimal(m.group(1))
            if n is not None: totals.append(n)
    total = max(totals) if totals else None
    return {
        "source": str(path.relative_to(RAW)),
        "document_type": kind,
        "number": parse_number(path, text),
        "date": parse_date(path, text),
        "counterparty_tins": counterparties,
        "total": str(total) if total is not None else "",
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    local = [d for p in sorted(RAW.rglob("*.xlsx")) if (d := parse_workbook(p))]
    esf = [r for r in json.loads(ESF.read_text(encoding="utf-8")) if r["active"] and r["direction"] == "outgoing"]
    links = []
    for doc in local:
        for inv in esf:
            if inv["buyer_tin"] not in doc["counterparty_tins"]:
                continue
            evidence = ["counterparty_tin"]
            score = 50
            if doc["total"] and doc["total"] == inv["total"]:
                score += 35; evidence.append("total")
            if doc["date"] and doc["date"] == inv["turnover_date"]:
                score += 15; evidence.append("event_date")
            if score >= 85:
                links.append({
                    "local_source": doc["source"], "local_type": doc["document_type"],
                    "local_number": doc["number"], "esf_registration_number": inv["registration_number"],
                    "esf_number": inv["number"], "score": score, "evidence": evidence,
                })
    payload = {"local_documents": local, "proposed_links": links}
    (OUT / "result.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {
        "local_documents": len(local),
        "with_counterparty_tin": sum(bool(d["counterparty_tins"]) for d in local),
        "with_total": sum(bool(d["total"]) for d in local),
        "active_outgoing_esf": len(esf),
        "proposed_links": len(links),
        "links_by_local_type": Counter(x["local_type"] for x in links),
    }
    (OUT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
