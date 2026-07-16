#!/usr/bin/env python3
"""Extract reviewable entity candidates from XML and Excel pilot documents."""

from __future__ import annotations

import csv
import json
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "working" / "entities"
OWN_TIN = "910226302322"


def clean(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def strip_ns(root: ET.Element) -> None:
    for el in root.iter():
        el.tag = el.tag.split("}")[-1]


def inner_xml_from_cdata(path: Path) -> list[ET.Element]:
    outer = ET.parse(path).getroot()
    result = []
    for body in outer.iter():
        if body.tag.split("}")[-1] == "invoiceBody" and body.text:
            inner = ET.fromstring(body.text.lstrip("\ufeff").strip())
            strip_ns(inner)
            result.append(inner)
    if not result:
        strip_ns(outer)
        result.append(outer)
    return result


def one(el: ET.Element, path: str) -> str:
    found = el.find(path)
    return clean(found.text) if found is not None else ""


def parse_xml(path: Path, counterparties: dict, items: dict, operations: list) -> None:
    roots = inner_xml_from_cdata(path)
    expanded = []
    for root in roots:
        awps = [el for el in root.iter() if el.tag.casefold() == "awp"]
        expanded.extend(awps or [root])
    for root in expanded:
        doc_type = "esf" if root.tag == "invoice" else "electronic_act" if "awp" in root.tag.casefold() or root.find("worksPerformed") is not None else "xml"
        doc_num = one(root, "num") or one(root, "number")
        doc_date = one(root, "date")
        turnover_date = one(root, "turnoverDate") or one(root, "performedDate")
        seller_tin = one(root, "./sellers/seller/tin") or one(root, "./senders/sender/tin")
        buyer_tin = one(root, "./customers/customer/tin") or one(root, "./recipients/recipient/tin")
        buyer_name = one(root, "./customers/customer/name") or one(root, "./recipients/recipient/name")
        seller_name = one(root, "./sellers/seller/name") or one(root, "./senders/sender/name")
        for tin, name, role in [(seller_tin, seller_name, "seller"), (buyer_tin, buyer_name, "buyer")]:
            if tin and tin != OWN_TIN:
                key = tin
                counterparties[key]["names"].add(name)
                counterparties[key]["sources"].add(str(path.relative_to(RAW)))
                counterparties[key]["roles"].add(role)
        total = one(root, "./productSet/totalPriceWithTax") or one(root, "./worksPerformed/totalSumWithTax")
        contract_num = one(root, "./deliveryTerm/contractNum") or one(root, "./contract/number")
        contract_date = one(root, "./deliveryTerm/contractDate") or one(root, "./contract/date")
        operations.append({
            "source": str(path.relative_to(RAW)), "document_type": doc_type, "number": doc_num,
            "document_date": doc_date, "event_date": turnover_date, "seller_tin": seller_tin,
            "buyer_tin": buyer_tin, "total": total, "contract_number": contract_num,
            "contract_date": contract_date,
        })
        for product in root.findall("./productSet/products/product"):
            name = one(product, "description") or one(product, "tnvedName")
            if name:
                key = name.casefold()
                items[key]["names"].add(name)
                items[key]["types"].add("goods")
                items[key]["sources"].add(str(path.relative_to(RAW)))
                items[key]["units"].add(one(product, "unitNomenclature"))
        for work in root.findall("./worksPerformed/works/work"):
            name = one(work, "name")
            if name:
                key = name.casefold()
                items[key]["names"].add(name)
                items[key]["types"].add("service")
                items[key]["sources"].add(str(path.relative_to(RAW)))
                items[key]["units"].add(one(work, "measureUnitCode"))


def parse_xlsx(path: Path, items: dict) -> None:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return
    for ws in wb.worksheets:
        rows = [[clean(c.value) for c in row] for row in ws.iter_rows()]
        header_idx = None
        name_col = None
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                v = val.casefold()
                if v == "наименование" or v.startswith("наименование, характеристика") or v.startswith("наименование работ"):
                    header_idx, name_col = i, j
                    break
            if header_idx is not None:
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1: header_idx + 35]:
            if name_col >= len(row):
                continue
            name = clean(row[name_col])
            if not name or name.casefold() in {"итого", "всего"} or len(name) < 3:
                continue
            if re.fullmatch(r"[0-9 xх.,-]+", name.casefold()):
                continue
            key = name.casefold()
            doc = str(path.relative_to(RAW)).casefold()
            kind = "service" if "авр" in doc or "акт выполн" in doc else "goods"
            items[key]["names"].add(name)
            items[key]["types"].add(kind)
            items[key]["sources"].add(str(path.relative_to(RAW)))
    wb.close()


def flatten_sets(record: dict) -> dict:
    return {k: sorted(v) if isinstance(v, set) else v for k, v in record.items()}


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    counterparties = defaultdict(lambda: {"names": set(), "roles": set(), "sources": set()})
    items = defaultdict(lambda: {"names": set(), "types": set(), "units": set(), "sources": set()})
    operations = []
    for path in sorted(RAW.rglob("*.xml")):
        parse_xml(path, counterparties, items, operations)
    for path in sorted(RAW.rglob("*.xlsx")):
        if not path.name.startswith("~$"):
            parse_xlsx(path, items)
    cp = [{"tin": tin, **flatten_sets(v)} for tin, v in counterparties.items()]
    ni = [{"normalized_key": key, **flatten_sets(v)} for key, v in items.items()]
    (OUT / "counterparties.json").write_text(json.dumps(cp, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "nomenclature_candidates.json").write_text(json.dumps(ni, ensure_ascii=False, indent=2), encoding="utf-8")
    with (OUT / "xml_operations.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(operations[0]) if operations else ["source"])
        w.writeheader(); w.writerows(operations)
    print(json.dumps({"counterparties_from_xml": len(cp), "nomenclature_candidates": len(ni), "xml_operations": len(operations)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
